import os
import re
import pandas as pd
from io import StringIO
from datetime import datetime
from pdfminer.high_level import extract_text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def extract_text_from_pdf(pdf_path):
    """Extrae el texto completo de un archivo PDF usando pdfminer.six"""
    try:
        text = extract_text(pdf_path)
        # Escribir el texto a un archivo para debug
        with open(f"{os.path.basename(pdf_path)}.txt", "w", encoding="utf-8") as f:
            f.write(text)
        return text
    except Exception as e:
        print(f"Error al extraer texto del PDF {pdf_path}: {e}")
        return ""

def parse_pdf_content(text_content, file_name):
    """Analiza el texto extraído de un PDF de resguardo y extrae los datos de los bienes."""
    records = []
    
    # Determinar el tipo de bien basado en el nombre del archivo
    default_tipo = 'Bien Mueble' if 'BM' in file_name else 'Enser Menor' if 'EM' in file_name else 'N/D'
    
    # Buscar patrones de información en el texto
    # Esta expresión busca líneas que tengan un ID (UTQ o EM seguido de números) seguido por una descripción
    lines = text_content.split("\n")
    
    # Imprimimos algunas líneas para debug
    print(f"Muestra de líneas del archivo {file_name}:")
    for i, line in enumerate(lines[:20]):
        if line.strip():
            print(f"  {i+1}: {line}")
    
    current_item = {}
    in_table = False
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Detectar inicio de tabla de datos
        if '"No.Act.' in line or '"EM' in line or '"UTQ' in line:
            in_table = True
            continue
            
        if in_table:
            # Intentar encontrar un ID (UTQ o EM seguido de números)
            id_match = re.search(r'"((?:UTQ|EM)\d+)"', line)
            
            if id_match:
                # Si ya teníamos un item en proceso, guardarlo
                if current_item and 'ID_Inventario' in current_item:
                    records.append(current_item)
                
                # Empezar un nuevo ítem
                current_item = {'ID_Inventario': id_match.group(1), 'Tipo_Bien': default_tipo}
                
                # Extraer el resto de la línea
                remaining = line[id_match.end():].strip()
                
                # Intentar extraer los campos separados por comas y entre comillas
                fields = re.findall(r'"([^"]*)"', remaining)
                
                if len(fields) >= 1:
                    current_item['Descripcion'] = fields[0].strip()
                if len(fields) >= 2:
                    current_item['No_Serie'] = fields[1].strip() or 'SIN NUMERO'
                if len(fields) >= 3:
                    current_item['Marca'] = fields[2].strip() or 'SIN MARCA'
                if len(fields) >= 4:
                    current_item['Modelo'] = fields[3].strip() or 'SIN MODELO'
                if len(fields) >= 5:
                    comentarios = fields[4].strip()
                    current_item['Observaciones'] = comentarios
                    
                    # Extraer datos de ubicación y estado del comentario
                    edificio = re.search(r'EDIFICIO:?\s*(?:"|["\s]*)?([^,"]+)(?:"|["\s]*)?', comentarios, re.IGNORECASE)
                    planta = re.search(r'PLANTA:?\s*(\w+)', comentarios, re.IGNORECASE)
                    area = re.search(r'(?:LABORATORIO|CUB[ÍI]CULO|ÁREA)[:\s]*([^,]+)', comentarios, re.IGNORECASE)
                    estado = re.search(r'(?:EDO\.DEL BIEN|ESTADO(?: DEL BIEN)?):?\s*(\w+)', comentarios, re.IGNORECASE)
                    
                    current_item['Edificio'] = edificio.group(1).strip().upper() if edificio else 'N/D'
                    current_item['Planta'] = planta.group(1).strip().upper() if planta else 'N/D'
                    current_item['Area_Especifica'] = area.group(1).strip() if area else 'N/D'
                    
                    estado_val = estado.group(1).strip().upper() if estado else 'BUENO'
                    # Normalizar Estado_Actual
                    if estado_val in ['BUENO', 'BUEN']:
                        estado_val = 'Bueno'
                    elif estado_val in ['REGULAR', 'REG']:
                        estado_val = 'Regular'
                    elif estado_val in ['MALO', 'MAL']:
                        estado_val = 'Malo'
                    elif 'REPARACION' in estado_val or 'REPARACIÓN' in estado_val:
                        estado_val = 'En Reparación'
                    elif 'BAJA' in estado_val:
                        estado_val = 'Dado de Baja'
                    else:
                        estado_val = 'Bueno'  # Valor por defecto
                    
                    current_item['Estado_Actual'] = estado_val
                    
            # Alternativa: puede ser una línea de continuación con datos adicionales
            elif current_item and 'ID_Inventario' in current_item:
                # Extraer más datos si hay comillas
                additional_fields = re.findall(r'"([^"]*)"', line)
                
                if additional_fields:
                    if 'Descripcion' not in current_item and len(additional_fields) > 0:
                        current_item['Descripcion'] = additional_fields[0].strip()
                    elif 'No_Serie' not in current_item and len(additional_fields) > 0:
                        current_item['No_Serie'] = additional_fields[0].strip() or 'SIN NUMERO'
                    elif 'Marca' not in current_item and len(additional_fields) > 0:
                        current_item['Marca'] = additional_fields[0].strip() or 'SIN MARCA'
                    elif 'Modelo' not in current_item and len(additional_fields) > 0:
                        current_item['Modelo'] = additional_fields[0].strip() or 'SIN MODELO'
                    elif 'Observaciones' not in current_item and len(additional_fields) > 0:
                        current_item['Observaciones'] = additional_fields[0].strip()
        
        # Si encontramos "Total de bienes" consideramos que terminó la tabla
        if 'Total de bienes' in line:
            in_table = False
            if current_item and 'ID_Inventario' in current_item:
                records.append(current_item)
                current_item = {}
    
    # No olvidar el último item si existe
    if current_item and 'ID_Inventario' in current_item:
        records.append(current_item)
    
    # Completar campos faltantes y normalizar
    for record in records:
        # Asegurar que todos los campos necesarios existan
        fields = ['ID_Inventario', 'Tipo_Bien', 'Descripcion', 'Marca', 'Modelo', 
                 'No_Serie', 'Edificio', 'Planta', 'Area_Especifica', 'Estado_Actual',
                 'Observaciones']
        
        for field in fields:
            if field not in record:
                if field == 'Tipo_Bien':
                    record[field] = default_tipo
                elif field in ['No_Serie']:
                    record[field] = 'SIN NUMERO'
                elif field in ['Marca']:
                    record[field] = 'SIN MARCA'
                elif field in ['Modelo']:
                    record[field] = 'SIN MODELO'
                elif field in ['Estado_Actual']:
                    record[field] = 'Bueno'
                else:
                    record[field] = ''
        
        # Normalizar nombres de edificios
        if 'Edificio' in record:
            edificio_val = record['Edificio']
            if 'NANO' in edificio_val: 
                record['Edificio'] = 'NANO'
            elif 'PIDET' in edificio_val: 
                record['Edificio'] = 'PIDET'
            elif 'CIC1' in edificio_val: 
                record['Edificio'] = 'CIC1'
        
        # Agregar campos de fechas
        record['Fecha_Ultima_Revision'] = datetime.now().strftime('%Y-%m-%d')
        record['Fecha_Resguardo_Original'] = ''
    
    return records

def create_excel_with_validations(df, output_file):
    """Crea un archivo Excel con validaciones de datos a partir del DataFrame"""
    # Guardar primero a CSV como respaldo
    csv_file = output_file.replace('.xlsx', '.csv')
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"Archivo CSV generado: {csv_file}")
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Definir las listas de validación
    tipos_bien = ['Bien Mueble', 'Enser Menor']
    edificios = sorted(list(set(df['Edificio'].dropna().unique().tolist() + ['NANO', 'PIDET', 'CIC1', 'CIC2'])))
    plantas = ['BAJA', 'ALTA', 'PRIMERA', 'SEGUNDA', 'TERCERA', 'CUARTA']
    estados = ['Bueno', 'Regular', 'Malo', 'En Reparación', 'Dado de Baja']
    
    # Agregar encabezados
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar ancho de columna según el encabezado
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 2)
    
    # Agregar datos
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            if pd.notna(value):  # Solo agregar valores no nulos
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Crear validaciones de datos
    # Para Tipo_Bien (columna B)
    tipo_dv = DataValidation(type="list", formula1=f'"{",".join(tipos_bien)}"', allow_blank=True)
    tipo_dv.add(f'B2:B{len(df) + 100}')  # Para filas adicionales
    ws.add_data_validation(tipo_dv)
    
    # Para Edificio (columna G)
    edificio_dv = DataValidation(type="list", formula1=f'"{",".join(edificios)}"', allow_blank=True)
    edificio_dv.add(f'G2:G{len(df) + 100}')
    ws.add_data_validation(edificio_dv)
    
    # Para Planta (columna H)
    planta_dv = DataValidation(type="list", formula1=f'"{",".join(plantas)}"', allow_blank=True)
    planta_dv.add(f'H2:H{len(df) + 100}')
    ws.add_data_validation(planta_dv)
    
    # Para Estado_Actual (columna J)
    estado_dv = DataValidation(type="list", formula1=f'"{",".join(estados)}"', allow_blank=True)
    estado_dv.add(f'J2:J{len(df) + 100}')
    ws.add_data_validation(estado_dv)
    
    # Formato para fechas (columnas K y M)
    for row in range(2, len(df) + 100):
        for col in [11, 13]:  # K es 11, M es 13 en números
            cell = ws.cell(row=row, column=col)
            cell.number_format = 'yyyy-mm-dd'
    
    # Guardar el archivo
    wb.save(output_file)
    print(f"Archivo Excel con validaciones generado: {output_file}")
    
    return output_file

def main():
    try:
        # Detectar PDFs en el directorio actual
        pdf_files = [f for f in os.listdir('.') if f.endswith('.pdf') and ('RESGUARDO' in f.upper() or 'INVENTARIO' in f.upper())]
        
        if not pdf_files:
            print("No se encontraron archivos PDF de resguardo en el directorio actual.")
            return
        
        print(f"Archivos PDF encontrados: {pdf_files}")
        all_data = []
        
        # Procesar cada PDF
        for pdf_file in pdf_files:
            print(f"Procesando {pdf_file}...")
            text = extract_text_from_pdf(pdf_file)
            if not text:
                print(f"  - No se pudo extraer texto del archivo {pdf_file}")
                continue
                
            records = parse_pdf_content(text, pdf_file)
            print(f"  - Se extrajeron {len(records)} registros.")
            all_data.extend(records)
        
        if not all_data:
            print("No se pudieron extraer datos de los PDFs.")
            return
        
        # Crear DataFrame y ordenar columnas
        df = pd.DataFrame(all_data)
        column_order = [
            'ID_Inventario', 'Tipo_Bien', 'Descripcion', 'Marca', 'Modelo',
            'No_Serie', 'Edificio', 'Planta', 'Area_Especifica', 'Estado_Actual',
            'Fecha_Ultima_Revision', 'Observaciones', 'Fecha_Resguardo_Original'
        ]
        
        # Asegurar que todas las columnas existan
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
                
        df = df[column_order]
        
        # Generar Excel con validaciones
        excel_file = "Control_Inventario.xlsx"
        create_excel_with_validations(df, excel_file)
        
        print("\nProceso completado exitosamente.")
        print(f"Se procesaron {len(all_data)} bienes en total.")
        print(f"Bienes Muebles: {len(df[df['Tipo_Bien'] == 'Bien Mueble'])}")
        print(f"Enseres Menores: {len(df[df['Tipo_Bien'] == 'Enser Menor'])}")
        
    except Exception as e:
        print(f"Error durante la ejecución: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 