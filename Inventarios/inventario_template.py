import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_excel_template(output_file, example_rows=5):
    """Crea una plantilla de Excel con la estructura requerida y validaciones"""
    # Crear un DataFrame vacío con la estructura requerida
    columns = [
        'ID_Inventario', 'Tipo_Bien', 'Descripcion', 'Marca', 'Modelo',
        'No_Serie', 'Edificio', 'Planta', 'Area_Especifica', 'Estado_Actual',
        'Fecha_Ultima_Revision', 'Observaciones', 'Fecha_Resguardo_Original'
    ]
    
    # Crear ejemplos de datos
    examples = []
    
    # Bien mueble ejemplo
    examples.append({
        'ID_Inventario': 'UTQ12345',
        'Tipo_Bien': 'Bien Mueble',
        'Descripcion': 'ESCRITORIO MADERA 4 CAJONES',
        'Marca': 'SIN MARCA',
        'Modelo': 'MADERA 4 CAJONES',
        'No_Serie': 'SIN NUMERO',
        'Edificio': 'NANO',
        'Planta': 'BAJA',
        'Area_Especifica': 'LABORATORIO 5',
        'Estado_Actual': 'Bueno',
        'Fecha_Ultima_Revision': datetime.now().strftime('%Y-%m-%d'),
        'Observaciones': '',
        'Fecha_Resguardo_Original': ''
    })
    
    # Enser menor ejemplo
    examples.append({
        'ID_Inventario': 'EM01234',
        'Tipo_Bien': 'Enser Menor',
        'Descripcion': 'SILLA',
        'Marca': 'SECRETARIAL',
        'Modelo': 'BASE DE ESTRELLA DE 5 PUNTAS',
        'No_Serie': 'SIN NUMERO',
        'Edificio': 'PIDET',
        'Planta': 'BAJA',
        'Area_Especifica': 'CUBÍCULO 3',
        'Estado_Actual': 'Bueno',
        'Fecha_Ultima_Revision': datetime.now().strftime('%Y-%m-%d'),
        'Observaciones': '',
        'Fecha_Resguardo_Original': ''
    })
    
    # Repetir ejemplos si se necesitan más filas
    while len(examples) < example_rows:
        examples.append({col: '' for col in columns})
    
    # Crear DataFrame
    df = pd.DataFrame(examples)
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Definir las listas de validación
    tipos_bien = ['Bien Mueble', 'Enser Menor']
    edificios = ['NANO', 'PIDET', 'CIC1', 'CIC2']
    plantas = ['BAJA', 'ALTA', 'PRIMERA', 'SEGUNDA', 'TERCERA', 'CUARTA']
    estados = ['Bueno', 'Regular', 'Malo', 'En Reparación', 'Dado de Baja']
    
    # Agregar encabezados con estilos
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Añadir borde
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        # Ajustar ancho de columna según el encabezado
        width = max(15, len(header) + 2)
        # Columnas específicas más anchas
        if header == 'Descripcion':
            width = 30
        elif header in ['Area_Especifica', 'Observaciones']:
            width = 25
        
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Agregar datos de ejemplo
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            if pd.notna(value):  # Solo agregar valores no nulos
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Establecer borde
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Centrar ciertas columnas
                if col_idx in [2, 7, 8, 10]:  # Tipo_Bien, Edificio, Planta, Estado_Actual
                    cell.alignment = Alignment(horizontal='center')
                
                # Color de fondo para filas ejemplo
                if row_idx <= 2:  # Primeros dos ejemplos
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Agregar filas adicionales vacías (total 100 filas)
    for row_idx in range(df.shape[0] + 2, 101):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            # Establecer borde
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Crear validaciones de datos
    # Para Tipo_Bien (columna B)
    tipo_dv = DataValidation(type="list", formula1=f'"{",".join(tipos_bien)}"', allow_blank=True)
    tipo_dv.add(f'B2:B101')
    ws.add_data_validation(tipo_dv)
    
    # Para Edificio (columna G)
    edificio_dv = DataValidation(type="list", formula1=f'"{",".join(edificios)}"', allow_blank=True)
    edificio_dv.add(f'G2:G101')
    ws.add_data_validation(edificio_dv)
    
    # Para Planta (columna H)
    planta_dv = DataValidation(type="list", formula1=f'"{",".join(plantas)}"', allow_blank=True)
    planta_dv.add(f'H2:H101')
    ws.add_data_validation(planta_dv)
    
    # Para Estado_Actual (columna J)
    estado_dv = DataValidation(type="list", formula1=f'"{",".join(estados)}"', allow_blank=True)
    estado_dv.add(f'J2:J101')
    ws.add_data_validation(estado_dv)
    
    # Formato para fechas (columnas K y M)
    for row in range(2, 101):
        for col in [11, 13]:  # K es 11, M es 13 en números
            cell = ws.cell(row=row, column=col)
            cell.number_format = 'yyyy-mm-dd'
    
    # Agregar una hoja de instrucciones
    ws_instructions = wb.create_sheet(title="Instrucciones")
    
    instructions = [
        ["INSTRUCCIONES PARA EL USO DE LA PLANTILLA DE INVENTARIO"],
        [""],
        ["1. Esta plantilla está diseñada para mantener un control organizado de bienes muebles y enseres menores."],
        [""],
        ["2. Estructura de Datos:"],
        ["   - ID_Inventario: Identificador único (No.Act. o EM). Campo obligatorio y único."],
        ["   - Tipo_Bien: Seleccione entre 'Bien Mueble' o 'Enser Menor' usando la lista desplegable."],
        ["   - Descripcion: Descripción detallada del bien."],
        ["   - Marca: Marca del bien (o 'SIN MARCA')."],
        ["   - Modelo: Modelo del bien (o 'SIN MODELO')."],
        ["   - No_Serie: Número de serie (o 'SIN NUMERO')."],
        ["   - Edificio: Seleccione el edificio de la lista desplegable."],
        ["   - Planta: Seleccione el nivel del edificio de la lista desplegable."],
        ["   - Area_Especifica: Ubicación detallada (ej: 'Laboratorio 5', 'Cubículo 3')."],
        ["   - Estado_Actual: Seleccione el estado del bien de la lista desplegable."],
        ["   - Fecha_Ultima_Revision: Fecha de la última verificación física del bien."],
        ["   - Observaciones: Campo libre para notas adicionales."],
        ["   - Fecha_Resguardo_Original: (Opcional) Fecha de asignación original."],
        [""],
        ["3. Las primeras dos filas contienen ejemplos de cómo llenar la información."],
        [""],
        ["4. Utilice las listas desplegables para mantener consistencia en los datos."],
        [""],
        ["5. Puede añadir más filas si es necesario, manteniendo el mismo formato."],
        [""],
        ["6. Para añadir edificios o valores adicionales a las listas desplegables, edite las validaciones de datos."]
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=row[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row[0].strip() and not row[0].startswith("   "):
            cell.font = Font(bold=True)
    
    # Ajustar ancho de columna
    ws_instructions.column_dimensions['A'].width = 100
    
    # Guardar el archivo
    wb.save(output_file)
    print(f"Plantilla Excel creada: {output_file}")
    
    return output_file

if __name__ == "__main__":
    excel_file = "Plantilla_Inventario.xlsx"
    create_excel_template(excel_file)
    print("\nSe ha creado una plantilla de Excel con la estructura solicitada.")
    print("La plantilla incluye ejemplos y validaciones de datos para mantener la consistencia.")
    print("Puede empezar a llenar sus datos directamente en el archivo Excel generado.") 