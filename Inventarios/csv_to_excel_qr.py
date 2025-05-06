import os
import pandas as pd
import qrcode
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image

def create_or_load_inventory(input_file=None, output_file="Control_Inventario_QR.xlsx"):
    """
    Crea un nuevo archivo de inventario o carga uno existente
    
    Args:
        input_file (str): Ruta al archivo CSV o Excel existente (opcional)
        output_file (str): Nombre del archivo Excel de salida
    
    Returns:
        DataFrame: DataFrame con los datos del inventario
    """
    required_columns = [
        'ID_Inventario', 'Tipo_Bien', 'Descripcion', 'Marca', 'Modelo',
        'No_Serie', 'Edificio', 'Planta', 'Area_Especifica', 'Estado_Actual',
        'Fecha_Ultima_Revision', 'Observaciones', 'Fecha_Resguardo_Original'
    ]
    
    if input_file and os.path.exists(input_file):
        # Cargar archivo existente
        print(f"Cargando archivo existente: {input_file}")
        if input_file.lower().endswith('.csv'):
            df = pd.read_csv(input_file)
        elif input_file.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(input_file)
        else:
            raise ValueError(f"Formato de archivo no soportado: {input_file}")
        
        # Verificar y añadir columnas faltantes
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
                print(f"Columna '{col}' no encontrada. Se ha creado vacía.")
        
        # Reordenar columnas
        df = df[required_columns]
        
    else:
        # Crear DataFrame nuevo con datos de ejemplo
        print("Creando nuevo archivo de inventario con datos de ejemplo...")
        
        # Datos de ejemplo (bienes muebles)
        example_data = [
            {
                'ID_Inventario': 'UTQ001',
                'Tipo_Bien': 'Bien Mueble',
                'Descripcion': 'ESCRITORIO EJECUTIVO CON 3 CAJONES',
                'Marca': 'SIN MARCA',
                'Modelo': 'EJECUTIVO MADERA',
                'No_Serie': 'SIN NUMERO',
                'Edificio': 'NANO',
                'Planta': 'BAJA',
                'Area_Especifica': 'OFICINA PRINCIPAL',
                'Estado_Actual': 'Bueno',
                'Fecha_Ultima_Revision': datetime.now().strftime('%Y-%m-%d'),
                'Observaciones': '',
                'Fecha_Resguardo_Original': '2020-01-15'
            },
            {
                'ID_Inventario': 'UTQ002',
                'Tipo_Bien': 'Bien Mueble',
                'Descripcion': 'SILLÓN EJECUTIVO ERGONÓMICO',
                'Marca': 'HERMAN MILLER',
                'Modelo': 'AERON',
                'No_Serie': 'HM234567',
                'Edificio': 'NANO',
                'Planta': 'BAJA',
                'Area_Especifica': 'OFICINA PRINCIPAL',
                'Estado_Actual': 'Bueno',
                'Fecha_Ultima_Revision': datetime.now().strftime('%Y-%m-%d'),
                'Observaciones': '',
                'Fecha_Resguardo_Original': '2020-01-15'
            },
            # Datos de ejemplo (enseres menores)
            {
                'ID_Inventario': 'EM001',
                'Tipo_Bien': 'Enser Menor',
                'Descripcion': 'ORGANIZADOR DE DOCUMENTOS',
                'Marca': 'OFFICE DEPOT',
                'Modelo': 'ESTÁNDAR 3 NIVELES',
                'No_Serie': 'SIN NUMERO',
                'Edificio': 'CIC1',
                'Planta': 'ALTA',
                'Area_Especifica': 'SALA DE JUNTAS',
                'Estado_Actual': 'Bueno',
                'Fecha_Ultima_Revision': datetime.now().strftime('%Y-%m-%d'),
                'Observaciones': '',
                'Fecha_Resguardo_Original': '2021-03-10'
            }
        ]
        
        df = pd.DataFrame(example_data)
    
    return df

def generate_qr_codes(df, qr_folder="codigos_qr"):
    """
    Genera códigos QR para cada elemento del inventario
    
    Args:
        df (DataFrame): DataFrame con los datos del inventario
        qr_folder (str): Carpeta donde se guardarán los códigos QR
    
    Returns:
        dict: Diccionario con las rutas a los archivos QR generados
    """
    print("Generando códigos QR...")
    
    # Crear carpeta para los códigos QR si no existe
    if not os.path.exists(qr_folder):
        os.makedirs(qr_folder)
        print(f"Carpeta para códigos QR creada: {qr_folder}")
    
    qr_paths = {}
    
    for idx, row in df.iterrows():
        id_inv = row['ID_Inventario']
        if not id_inv:
            continue
            
        # Crear un diccionario con la información del bien
        item_info = {
            'ID': id_inv,
            'Tipo': row['Tipo_Bien'],
            'Descripción': row['Descripcion'],
            'Marca': row['Marca'],
            'Modelo': row['Modelo'],
            'No_Serie': row['No_Serie'],
            'Ubicación': f"{row['Edificio']}, {row['Planta']}, {row['Area_Especifica']}",
            'Estado': row['Estado_Actual']
        }
        
        # Crear el código QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        # Convertir el diccionario a texto para el QR
        qr_text = "\n".join([f"{k}: {v}" for k, v in item_info.items() if v])
        qr.add_data(qr_text)
        qr.make(fit=True)
        
        # Crear la imagen del QR
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Guardar la imagen
        qr_path = os.path.join(qr_folder, f"{id_inv}.png")
        img.save(qr_path)
        qr_paths[id_inv] = qr_path
    
    print(f"Se generaron {len(qr_paths)} códigos QR en la carpeta: {qr_folder}")
    
    return qr_paths

def create_excel_with_qr(df, qr_paths, excel_output="Control_Inventario_QR.xlsx"):
    """
    Crea un archivo Excel con validaciones y enlaces a códigos QR
    
    Args:
        df (DataFrame): DataFrame con los datos del inventario
        qr_paths (dict): Diccionario con las rutas a los archivos QR
        excel_output (str): Nombre del archivo Excel de salida
    
    Returns:
        str: Ruta al archivo Excel generado
    """
    print("Creando archivo Excel con validaciones y códigos QR...")
    
    required_columns = [
        'ID_Inventario', 'Tipo_Bien', 'Descripcion', 'Marca', 'Modelo',
        'No_Serie', 'Edificio', 'Planta', 'Area_Especifica', 'Estado_Actual',
        'Fecha_Ultima_Revision', 'Observaciones', 'Fecha_Resguardo_Original'
    ]
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Definir las listas de validación
    tipos_bien = ['Bien Mueble', 'Enser Menor']
    edificios = sorted(list(set(filter(None, df['Edificio'].dropna().unique().tolist() + ['NANO', 'PIDET', 'CIC1', 'CIC2']))))
    plantas = ['BAJA', 'ALTA', 'PRIMERA', 'SEGUNDA', 'TERCERA', 'CUARTA']
    estados = ['Bueno', 'Regular', 'Malo', 'En Reparación', 'Dado de Baja']
    
    # Agregar encabezados con estilos
    headers = required_columns + ["Código QR"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Añadir borde
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        # Ajustar ancho de columna según el encabezado
        width = max(15, len(header) + 2)
        # Columnas específicas más anchas
        if header == 'Descripcion':
            width = 30
        elif header in ['Area_Especifica', 'Observaciones']:
            width = 25
        elif header == 'Código QR':
            width = 20
            
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Ajustar altura de filas
    ws.row_dimensions[1].height = 30
    
    # Agregar datos
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            if pd.notna(value):  # Solo agregar valores no nulos
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Establecer borde
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Centrar ciertas columnas
                if col_idx in [2, 7, 8, 10]:  # Tipo_Bien, Edificio, Planta, Estado_Actual
                    cell.alignment = Alignment(horizontal='center')
        
        # Agregar enlace al código QR
        id_inv = row[0]  # ID_Inventario es la primera columna
        if id_inv and id_inv in qr_paths:
            qr_link_cell = ws.cell(row=row_idx, column=len(headers))
            qr_link_cell.value = f"Código QR {id_inv}"
            qr_link_cell.hyperlink = qr_paths[id_inv]
            qr_link_cell.style = "Hyperlink"
            qr_link_cell.alignment = Alignment(horizontal='center')
            qr_link_cell.border = thin_border
    
    # Ajustar altura de filas para los datos
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 22
    
    # Crear validaciones de datos
    # Para Tipo_Bien (columna B)
    tipo_dv = DataValidation(type="list", formula1=f'"{",".join(tipos_bien)}"', allow_blank=True)
    tipo_dv.add(f'B2:B{len(df) + 100}')
    ws.add_data_validation(tipo_dv)
    
    # Para Edificio (columna G)
    edificio_dv = DataValidation(type="list", formula1=f'"{",".join(edificios)}"', allow_blank=True)
    edificio_dv.add(f'G2:G{len(df) + 100}')
    ws.add_data_validation(edificio_dv)
    
    # Para Planta (columna H)
    planta_dv = DataValidation(type="list", formula1=f'"{",".join(plantas)}"', allow_blank=True)
    planta_dv.add(f'H2:H{len(df) + 100}')
    ws.add_data_validation(planta_dv)
    
    # Para Estado_Actual (columna J)
    estado_dv = DataValidation(type="list", formula1=f'"{",".join(estados)}"', allow_blank=True)
    estado_dv.add(f'J2:J{len(df) + 100}')
    ws.add_data_validation(estado_dv)
    
    # Formato para fechas (columnas K y M)
    for row in range(2, len(df) + 100):
        for col in [11, 13]:  # K es 11, M es 13 en números
            cell = ws.cell(row=row, column=col)
            cell.number_format = 'yyyy-mm-dd'
    
    # Crear hoja adicional para vista de QR
    ws_qr = wb.create_sheet(title="Códigos QR")
    
    # Agregar cabecera
    ws_qr.cell(row=1, column=1, value="ID_Inventario").font = Font(bold=True)
    ws_qr.cell(row=1, column=2, value="Descripcion").font = Font(bold=True)
    ws_qr.cell(row=1, column=3, value="Código QR").font = Font(bold=True)
    
    # Ajustar anchos de columnas
    ws_qr.column_dimensions['A'].width = 15
    ws_qr.column_dimensions['B'].width = 30
    ws_qr.column_dimensions['C'].width = 25
    
    # Añadir altura a las filas
    row_height = 150  # Altura para filas con imágenes QR
    
    # Agregar registros con códigos QR
    qr_row = 2
    for idx, row in df.iterrows():
        id_inv = row['ID_Inventario']
        if not id_inv or id_inv not in qr_paths:
            continue
            
        ws_qr.cell(row=qr_row, column=1, value=id_inv)
        ws_qr.cell(row=qr_row, column=2, value=row['Descripcion'])
        
        # Insertar imagen QR
        try:
            img = Image(qr_paths[id_inv])
            # Redimensionar la imagen para que quepa bien en la celda
            img.width = 120
            img.height = 120
            ws_qr.add_image(img, f'C{qr_row}')
            
            # Ajustar altura de fila
            ws_qr.row_dimensions[qr_row].height = row_height
            qr_row += 1
        except Exception as e:
            print(f"Error al añadir código QR para {id_inv}: {e}")
            ws_qr.cell(row=qr_row, column=3, value=f"Ver archivo: {qr_paths[id_inv]}")
            qr_row += 1
    
    # Crear hoja para impresión de etiquetas
    ws_print = wb.create_sheet(title="Etiquetas QR")
    
    # Configurar hoja para impresión
    ws_print.page_setup.orientation = ws_print.ORIENTATION_PORTRAIT
    ws_print.page_setup.paperSize = ws_print.PAPERSIZE_A4
    
    # Ajustar márgenes
    margin = 0.5  # en pulgadas
    ws_print.page_margins.left = margin
    ws_print.page_margins.right = margin
    ws_print.page_margins.top = margin
    ws_print.page_margins.bottom = margin
    
    # Configuración para 2 columnas x 5 filas (10 QR por página)
    cols = 2
    rows_per_page = 5
    
    # Dimensionar columnas y filas
    for col in range(1, cols*2 + 1):
        ws_print.column_dimensions[get_column_letter(col)].width = 12
    
    # Añadir QR a la hoja
    current_row = 1
    current_col = 1
    items_added = 0
    
    for id_inv, qr_path in qr_paths.items():
        item = df[df['ID_Inventario'] == id_inv].iloc[0] if id_inv in df['ID_Inventario'].values else None
        
        if item is None:
            continue
        
        # Ajustar altura de fila
        ws_print.row_dimensions[current_row].height = 110
        
        # Agregar ID y descripción
        ws_print.cell(row=current_row, column=current_col).value = id_inv
        ws_print.cell(row=current_row, column=current_col).font = Font(bold=True)
        ws_print.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center')
        
        desc_cell = ws_print.cell(row=current_row, column=current_col + 1)
        desc_cell.value = item['Descripcion'][:30] + "..." if len(item['Descripcion']) > 30 else item['Descripcion']
        desc_cell.alignment = Alignment(wrap_text=True)
        
        # Agregar imagen QR
        try:
            img = Image(qr_path)
            # Ajustar tamaño
            img.width = 100
            img.height = 100
            # Posicionar debajo del texto
            ws_print.add_image(img, f'{get_column_letter(current_col)}{current_row + 1}')
        except Exception as e:
            ws_print.cell(row=current_row + 1, column=current_col).value = "Error QR"
        
        # Mover a la siguiente posición
        current_col += 2
        if current_col > cols*2:
            current_col = 1
            current_row += 3  # Espacio para QR + texto + margen
        
        items_added += 1
        
        # Añadir salto de página cada 'rows_per_page' filas
        if items_added % (cols * rows_per_page) == 0 and items_added < len(qr_paths):
            ws_print.page_setup.fitToPage = True
            ws_print.row_dimensions[current_row].height = 10
            ws_print.cell(row=current_row, column=1).value = "--- Cortar aquí ---"
            current_row += 1
    
    # Crear hoja de instrucciones
    ws_instructions = wb.create_sheet(title="Instrucciones")
    
    instructions = [
        ["INSTRUCCIONES PARA EL USO DEL CONTROL DE INVENTARIO"],
        [""],
        ["1. Este archivo permite gestionar el inventario de bienes muebles y enseres menores con validaciones de datos y códigos QR."],
        [""],
        ["2. Hoja 'Inventario':"],
        ["   - Contiene todos los bienes inventariados con sus características."],
        ["   - Se han configurado validaciones para mantener la consistencia de los datos."],
        ["   - La columna 'Código QR' contiene enlaces a los archivos QR generados."],
        [""],
        ["3. Hoja 'Códigos QR':"],
        ["   - Muestra los códigos QR generados para cada ítem del inventario."],
        ["   - Puede usar esta hoja para visualizar los códigos QR generados."],
        [""],
        ["4. Hoja 'Etiquetas QR':"],
        ["   - Diseñada para imprimir etiquetas con códigos QR (2 columnas x 5 filas por página)."],
        ["   - Optimizada para impresión en papel adhesivo o normal para recortar."],
        [""],
        ["5. Directorio 'codigos_qr':"],
        ["   - Contiene los archivos individuales de cada código QR."],
        ["   - Puede utilizarlos para impresión, etiquetado o generación de reportes."],
        [""],
        ["6. Para agregar nuevos elementos al inventario:"],
        ["   - Añada una nueva fila en la hoja 'Inventario'."],
        ["   - Utilice las listas desplegables para garantizar consistencia."],
        ["   - Ejecute nuevamente el script para generar los códigos QR de los nuevos elementos."],
        [""],
        ["7. Cada código QR contiene la información esencial del bien:"],
        ["   - ID, tipo, descripción, marca, modelo, número de serie, ubicación y estado."],
        ["   - Se puede escanear con cualquier aplicación de lectura de códigos QR."],
        [""],
        ["8. Consideraciones:"],
        ["   - Realice copias de seguridad periódicas de este archivo."],
        ["   - Para modificaciones masivas, es recomendable exportar a CSV, editar y volver a procesar."]
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=row[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row[0].strip() and not row[0].startswith("   "):
            cell.font = Font(bold=True)
    
    # Ajustar ancho de columna
    ws_instructions.column_dimensions['A'].width = 120
    
    # Guardar el archivo Excel
    wb.save(excel_output)
    print(f"Archivo Excel creado exitosamente: {excel_output}")
    
    return excel_output

def main():
    """Función principal"""
    print("Control de Inventario con Códigos QR")
    print("===================================")
    
    # Buscar archivos CSV o Excel en el directorio actual
    csv_files = [f for f in os.listdir('.') if f.lower().endswith('.csv')]
    excel_files = [f for f in os.listdir('.') if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    input_file = None
    
    if csv_files or excel_files:
        print("Archivos encontrados:")
        
        all_files = []
        file_index = 1
        
        # Mostrar archivos CSV
        for csv in csv_files:
            print(f"  {file_index}. {csv} (CSV)")
            all_files.append(csv)
            file_index += 1
        
        # Mostrar archivos Excel
        for excel in excel_files:
            print(f"  {file_index}. {excel} (Excel)")
            all_files.append(excel)
            file_index += 1
        
        # Opción para crear nuevo
        print(f"  {file_index}. Crear nuevo archivo de inventario con datos de ejemplo")
        
        try:
            choice = int(input("\nSeleccione una opción: "))
            if 1 <= choice < file_index:
                input_file = all_files[choice - 1]
        except (ValueError, IndexError):
            print("Opción no válida. Se creará un nuevo archivo de inventario.")
            input_file = None
    else:
        print("No se encontraron archivos CSV o Excel en el directorio actual.")
        print("Se creará un nuevo archivo de inventario con datos de ejemplo.")
    
    try:
        # Cargar o crear datos de inventario
        df = create_or_load_inventory(input_file)
        
        # Generar códigos QR
        qr_paths = generate_qr_codes(df)
        
        # Crear Excel con códigos QR
        excel_file = create_excel_with_qr(df, qr_paths)
        
        print("\n=== Proceso completado ===")
        print(f"1. Excel con inventario: {excel_file}")
        print(f"2. Carpeta con códigos QR individuales: codigos_qr/\n")
        
        print("Instrucciones:")
        print("- Use el archivo Excel generado para gestionar su inventario.")
        print("- Utilice la hoja 'Etiquetas QR' para imprimir los códigos QR.")
        print("- Escanee los códigos QR con cualquier lector de QR para ver la información del bien.")
        
    except Exception as e:
        print(f"Error durante la ejecución: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 